#!/usr/bin/env python3
"""
================================================================================
GÉNÉRATEUR DE PLANNING - VERSION 35 (FIX MISSING FUNCTIONS)
================================================================================

CORRECTIF URGENT :
- Réintégration des fonctions 'get_excel_ranges_for_period' et
  'get_excel_ranges_for_ferie' qui manquaient dans la V34.

FONCTIONNALITÉS V34 MAINTENUES :
- Config : AG (Grève) en rouge, Aptos 11 Centré.
- Dashboard : Colonne A largeur 5.26cm.
- Mensuel : Bordures jours 29-31, Total Heures, AG rouge.

Usage:
    python generer_planning_v35.py 2027
================================================================================
"""

import sys
import os
from datetime import datetime, timedelta
import calendar as cal
import copy

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import CellIsRule
except ImportError:
    print("\n❌ Module openpyxl requis")
    print("Installation: pip install openpyxl --break-system-packages\n")
    sys.exit(1)

# --- CONFIGURATION ---
COLOR_CV_WATER_GREEN = '96D1CC'
COLOR_AG_RED = 'FF4444'

COLORS = {
    '101': 'FFD9E6', '6101': 'FFD9E6', '7101': 'FFD9E6',
    '102': 'FFE6F0', '6102': 'FFE6F0', '7102': 'FFE6F0',
    '111': 'D9E6FF', '6111': 'D9E6FF', '7111': 'D9E6FF',
    '112': 'E6F0FF', '6112': 'E6F0FF', '7112': 'E6F0FF',
    '121': 'FFE6CC', '6121': 'FFE6CC', '7121': 'FFE6CC',
    'X_AM': 'D9F2D9', 'X_PM': 'D9F2D9',
    'X_10': 'E6D9FF',
    'CN': 'FFFFCC', 'JC': 'FFFFCC',
    'RH': 'CCCCCC', 'CH': 'D5D5D5',
    'RR': 'C9C9C9', 'ZM': 'F0F0F0',
    'CV': COLOR_CV_WATER_GREEN,
    'AG': COLOR_AG_RED
}

LISTE_PRESTATIONS = ['101', '102', '111', '112', '121', '6101', '6102', '6111', '6112', '6121', '7101', '7102', '7111', '7112', '7121', 'X_AM', 'X_PM', 'X_10', 'AG']
LISTE_REPOS = ['CN', 'JC', 'RH', 'CH', 'RR', 'CV', 'ZM']
CODES_8H = ['101', '102', '111', '112', '121', '6101', '6102', '6111', '6112', '6121',
            '7101', '7102', '7111', '7112', '7121', 'X_AM', 'X_PM', 'X_10', 'AG',
            'CN', 'JC', 'RR', 'ZM', 'CV']
TOUS_LES_CODES = LISTE_PRESTATIONS + LISTE_REPOS
LISTE_DASHBOARD = LISTE_PRESTATIONS + LISTE_REPOS

mois_noms = ['Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin', 'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre']

# --- LOGIQUE HEURES DE NUIT ---
CODES_NUIT_8H = ['121', '6121', '7121']
CODES_NUIT_2H = ['101', '102', '6101', '6102', '7101', '7102', 'X_AM',
                 '111', '112', '6111', '6112', '7111', '7112', 'X_PM']

def get_night_hours_formula(code, cell_count_ref):
    str_code = str(code)
    if str_code in CODES_NUIT_8H: return f'={cell_count_ref}*8'
    elif str_code in CODES_NUIT_2H: return f'={cell_count_ref}*2'
    return 0

# --- FONCTION NETTOYAGE ---
def clean_zone_restricted(ws, row_start, row_end, col_start, col_end):
    # Défusion
    ranges_to_unmerge = []
    for mr in list(ws.merged_cells.ranges):
        if (mr.min_row >= row_start and mr.max_row <= row_end and
            mr.min_col >= col_start and mr.max_col <= col_end):
            ranges_to_unmerge.append(mr)
    for mr in ranges_to_unmerge:
        try: ws.unmerge_cells(str(mr))
        except: pass

    # Vider
    for r in range(row_start, row_end + 1):
        for c in range(col_start, col_end + 1):
            try:
                cell = ws.cell(row=r, column=c)
                cell.value = None
                cell.fill = PatternFill(fill_type=None)
                cell.border = Border()
                cell.font = Font()
                cell.number_format = 'General'
            except: pass

# --- FONCTIONS CROSS-SHEET (RÉINTÉGRÉES) ---
def get_excel_ranges_for_period(start_date, end_date, annee):
    """Calcule les plages Excel pour une période donnée (Cross-Onglets)."""
    ranges = []
    current_date = start_date
    while current_date <= end_date:
        _, last_day = cal.monthrange(current_date.year, current_date.month)
        end_m = datetime(current_date.year, current_date.month, last_day)
        seg_end = min(end_date, end_m)
        s_name = mois_noms[current_date.month - 1]

        d1 = max(current_date, datetime(current_date.year, current_date.month, 1))
        d2 = min(seg_end, datetime(current_date.year, current_date.month, 15))
        if d1 <= d2:
            ranges.append(f"'{s_name}'!{get_column_letter(d1.day+1)}6:{get_column_letter(d2.day+1)}6")

        if last_day > 15:
            d3 = max(current_date, datetime(current_date.year, current_date.month, 16))
            d4 = seg_end
            if d3 <= d4:
                ranges.append(f"'{s_name}'!{get_column_letter(d3.day-14)}12:{get_column_letter(d4.day-14)}12")
        current_date = end_m + timedelta(days=1)
    return ranges

def get_excel_ranges_for_ferie(start_date, end_date, annee, feries):
    """Retourne les cellules spécifiques fériées."""
    cells = []
    curr = start_date
    while curr <= end_date:
        if curr in feries:
            s_name = mois_noms[curr.month - 1]
            if curr.day <= 15: c, r = get_column_letter(curr.day+1), 6
            else: c, r = get_column_letter(curr.day-14), 12
            cells.append(f"'{s_name}'!{c}{r}")
        curr += timedelta(days=1)
    return cells

# --- INITIALISATION ---
if len(sys.argv) != 2:
    print("\nUsage: python generer_planning_v35.py <ANNEE>")
    sys.exit(1)

try:
    ANNEE = int(sys.argv[1])
except ValueError:
    print("❌ L'année doit être un nombre")
    sys.exit(1)

TEMPLATE = "planning_2026_FINAL.xlsx"
if not os.path.exists(TEMPLATE):
    print(f"❌ {TEMPLATE} introuvable.")
    sys.exit(1)

print("\n" + "="*70)
print(f"🗓️  GÉNÉRATION PLANNING {ANNEE}")
print("="*70)

def calculer_periodes_dynamiques(annee_cible):
    anchor_start = datetime(2026, 1, 12)
    delta_years = annee_cible - 2026
    start_target = anchor_start + timedelta(days=364 * delta_years)
    periodes = []
    for i in range(13):
        p_start = start_target + timedelta(days=28 * i)
        p_end = p_start + timedelta(days=27)
        periodes.append((i+1, p_start, p_end))
    return periodes

def get_jours_feries(annee):
    feries = {}
    for y in [annee, annee+1]:
        a, b, c = y % 19, y // 100, y % 100
        d, e = b // 4, b % 4
        f = (b + 8) // 25
        g = (b - f + 1) // 3
        h = (19 * a + b - d - g + 15) % 30
        i, k = c // 4, c % 4
        l = (32 + 2 * e + 2 * i - h - k) % 7
        m = (a + 11 * h + 22 * l) // 451
        mois = (h + l - 7 * m + 114) // 31
        jour = ((h + l - 7 * m + 114) % 31) + 1
        paques = datetime(y, mois, jour)
        feries.update({
            datetime(y, 1, 1): "Nouvel An",
            paques + timedelta(days=1): "Lundi Pâques",
            datetime(y, 5, 1): "Fête Travail",
            paques + timedelta(days=39): "Ascension",
            paques + timedelta(days=50): "Pentecôte",
            datetime(y, 7, 21): "Fête Nat.",
            datetime(y, 8, 15): "Assomption",
            datetime(y, 11, 1): "Toussaint",
            datetime(y, 11, 11): "Armistice",
            datetime(y, 12, 25): "Noël"
        })
    return feries

def get_periode_pour_date(date, periodes):
    for num, debut, fin in periodes:
        if debut <= date <= fin: return num, debut, fin
    return None, None, None

jours_feries = get_jours_feries(ANNEE)
periodes = calculer_periodes_dynamiques(ANNEE)
jours_semaine_abbr = ['L', 'M', 'M', 'J', 'V', 'S', 'D']

print(f"\n📂 Chargement {TEMPLATE}...")
wb = load_workbook(TEMPLATE)

# Styles
border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
fill_header = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
font_header = Font(name='Aptos', bold=True, color="FFFFFF", size=10)
fill_total = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
font_total = Font(name='Aptos', bold=True, color="000000", size=10)
fill_none = PatternFill(fill_type=None)
font_data = Font(name='Aptos', bold=False, color="000000", size=10)
centrage = Alignment(horizontal='center', vertical='center')

fill_weekend = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
fill_ferie = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
fill_period_bound = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
border_thick_left = Border(left=Side(style='thick', color='0070C0'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
border_thick_right = Border(left=Side(style='thin'), right=Side(style='thick', color='0070C0'), top=Side(style='thin'), bottom=Side(style='thin'))
fill_cv = PatternFill(start_color=COLOR_CV_WATER_GREEN, end_color=COLOR_CV_WATER_GREEN, fill_type="solid")
fill_ag = PatternFill(start_color=COLOR_AG_RED, end_color=COLOR_AG_RED, fill_type="solid")

alert_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
alert_font = Font(name='Aptos', bold=True, color="FFFFFF", size=10)
warning_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

# =============================================================================
# 1. CONFIGURATION
# =============================================================================
print("  ✓ Mise à jour 'CONFIGURATION'...")
config = wb['CONFIGURATION']
config['A1'].value = f'CONFIGURATION DU FICHIER - ANNÉE {ANNEE}'
config['B4'].value = ANNEE

row_p_start = 8
for r in range(1, 20):
    val = config[f'A{r}'].value
    if val and 'P1' in str(val) and len(str(val)) == 2:
        row_p_start = r
        break

# Écriture Dates uniquement
for i, (num, start, end) in enumerate(periodes):
    r = row_p_start + i
    config[f'A{r}'].value = f'P{num}'
    config[f'B{r}'].value = start
    config[f'C{r}'].value = end
    config[f'D{r}'].value = 160
    config[f'B{r}'].number_format = 'dd/mm/yyyy'
    config[f'C{r}'].number_format = 'dd/mm/yyyy'
    for c_idx in range(1, 5):
        cell = config.cell(row=r, column=c_idx)
        cell.border = border_thin
        cell.alignment = centrage
        cell.font = Font(name='Aptos', size=10)

# AJOUT AG LIGNE 57
config['A57'].value = "AG"
config['B57'].value = "Grève"
config['A57'].fill = fill_ag
config['A57'].font = Font(name='Aptos', bold=True, color="000000")
config['A57'].alignment = centrage
config['A57'].border = border_thin

# Style spécifique demandé pour la description "Grève"
config['B57'].font = Font(name='Aptos', size=11)
config['B57'].alignment = centrage
config['B57'].border = border_thin

row_cn_conf = 6
row_jc_conf = 7
for r in range(1, 15):
    val = str(config[f'A{r}'].value).strip()
    if val == 'CN': row_cn_conf = r
    if val == 'JC': row_jc_conf = r

# =============================================================================
# 2. TABLEAU DE BORD
# =============================================================================
print("  ✓ Traitement 'Tableau de Bord'...")
dash = wb['TABLEAU DE BORD']
dash['A1'].value = f'TABLEAU DE BORD - ANNÉE {ANNEE}'

# Largeur colonne A : ~5.26cm (environ 26 caractères Excel)
dash.column_dimensions['A'].width = 26

clean_zone_restricted(dash, 12, 100, 1, 5)

curr = 12
row_cn_stats = 0
row_jc_stats = 0

for code in LISTE_DASHBOARD:
    if code == 'CN': row_cn_stats = curr
    if code == 'JC': row_jc_stats = curr

    dash[f'A{curr}'].value = code
    parts = []
    for m in mois_noms:
        parts.append(f'COUNTIF({m}!B6:P6,"{code}")')
        parts.append(f'COUNTIF({m}!B12:R12,"{code}")')
    dash[f'B{curr}'].value = "=" + "+".join(parts)

    if code in CODES_8H: dash[f'C{curr}'].value = f'=B{curr}*8'
    else: dash[f'C{curr}'].value = 0

    dash[f'D{curr}'].value = get_night_hours_formula(code, f'B{curr}')

    for col in ['A','B','C','D']:
        c = dash[f'{col}{curr}']
        c.border = border_thin
        c.alignment = centrage
        c.fill = fill_none
        if col in ['C', 'D']: c.number_format = 'General'

        if code == 'CV':
            c.fill = fill_cv
            c.font = Font(name='Aptos', size=11, color="000000")
        elif code == 'AG':
            c.fill = fill_ag
            c.font = Font(name='Aptos', size=11, color="000000")
        else:
            c.font = Font(name='Aptos', size=11, color="000000")
    curr += 1

# TOTAL
dash[f'A{curr}'].value = "TOTAL"
dash[f'C{curr}'].value = f'=SUM(C{12}:C{curr-1})'
dash[f'D{curr}'].value = f'=SUM(D{12}:D{curr-1})'
for col in ['A','B','C','D']:
    c = dash[f'{col}{curr}']
    c.border = border_thin
    c.alignment = centrage
    c.fill = fill_total
    c.font = Font(name='Aptos', size=11, bold=True, color="000000")
    if col in ['C', 'D']: c.number_format = 'General'
curr += 2

# REPORT
dash[f'A{curr}'].value = f"REPORT POUR {ANNEE+1}"
dash[f'A{curr}'].font = Font(bold=True, size=11)
curr += 1
headers_report = ['Type', 'Droit', 'Report', 'Pris', 'Solde']
for idx, h in enumerate(headers_report):
    c = dash.cell(row=curr, column=idx+1)
    c.value = h
    c.fill = fill_header
    c.font = font_header
    c.alignment = centrage
    c.border = border_thin
curr += 1

configs = {'CN': (row_cn_conf, row_cn_stats), 'JC': (row_jc_conf, row_jc_stats)}
for code, (r_conf, r_stats) in configs.items():
    dash[f'A{curr}'].value = code
    dash[f'B{curr}'].value = f'=CONFIGURATION!B{r_conf}'
    dash[f'C{curr}'].value = f'=CONFIGURATION!C{r_conf}'
    if r_stats > 0: dash[f'D{curr}'].value = f'=B{r_stats}'
    else: dash[f'D{curr}'].value = 0
    dash[f'E{curr}'].value = f'=B{curr}+C{curr}-D{curr}'
    for col in range(1, 6):
        c = dash.cell(row=curr, column=col)
        c.border = border_thin
        c.font = Font(name='Aptos', size=10, color="000000")
        c.alignment = centrage
        c.number_format = 'General'
    curr += 1

dash['E6'].value = f'=B{row_cn_stats}' if row_cn_stats else 0
dash['E7'].value = f'=B{row_jc_stats}' if row_jc_stats else 0
dash['F6'].value = '=B6+C6-E6'
dash['F7'].value = '=B7+C7-E7'

# =============================================================================
# 3. TRAITEMENT MENSUEL
# =============================================================================
print("  ✓ Génération des calendriers mensuels...")

for mois_num in range(1, 13):
    mois_nom = mois_noms[mois_num - 1]
    ws = wb[mois_nom]
    ws['A1'].value = f'{mois_nom.upper()} {ANNEE}'
    ws.column_dimensions['A'].width = 45

    _, nb_jours = cal.monthrange(ANNEE, mois_num)

    # 1. Calendrier
    for j in range(29, 32):
        col = get_column_letter(j-14)
        for r in [10, 11, 12, 13]:
            try:
                ws[f'{col}{r}'].value = None
                ws[f'{col}{r}'].fill = fill_none
                ws[f'{col}{r}'].border = Border()
            except: pass

    for j in range(1, 32):
        if j > nb_jours: continue
        date = datetime(ANNEE, mois_num, j)
        idx = date.weekday()
        if j <= 15: r_n, r_l, r_s, r_p = 4, 5, 6, 7; c_idx = j+1
        else: r_n, r_l, r_s, r_p = 10, 11, 12, 13; c_idx = j-14
        col = get_column_letter(c_idx)

        c_num, c_let, c_saisie, c_per = ws[f'{col}{r_n}'], ws[f'{col}{r_l}'], ws[f'{col}{r_s}'], ws[f'{col}{r_p}']
        c_num.value = j
        c_let.value = jours_semaine_abbr[idx]

        for c in [c_num, c_let, c_saisie]:
            c.fill = fill_none
            c.font = Font(name='Aptos', size=10)

        # FIX BORDURES : Réapplication explicite de la bordure sur la case de saisie
        c_saisie.border = border_thin

        num_p, deb_p, fin_p = get_periode_pour_date(date, periodes)
        if num_p:
            c_per.value = f'P{num_p}'
            is_bound = date in [deb_p, fin_p]
            c_per.font = Font(name='Aptos', size=8, bold=is_bound, color="0070C0" if is_bound else "000000")
            if date == deb_p:
                for c in [c_num, c_let, c_per]: c.fill = fill_period_bound
                c_saisie.border = border_thick_left
            elif date == fin_p:
                for c in [c_num, c_let, c_per]: c.fill = fill_period_bound
                c_saisie.border = border_thick_right
            else:
                c_per.fill = fill_none
                c_per.border = Border()

        is_weekend = idx in [5, 6]
        is_ferie = date in jours_feries
        if is_weekend and not (num_p and date in [deb_p, fin_p]):
            c_num.fill = fill_weekend
            c_let.fill = fill_weekend
        if is_ferie:
            c_num.fill = fill_ferie
            c_let.fill = fill_ferie
            c_num.font = Font(name='Aptos', size=10, bold=True)

    # 2. Stats Mensuelles
    ws.data_validations.dataValidation = []
    str_codes = ",".join(TOUS_LES_CODES)
    dv = DataValidation(type="list", formula1=f'"{str_codes}"', allow_blank=True)
    ws.add_data_validation(dv)
    rng1 = f'B6:{get_column_letter(min(16, nb_jours)+1)}6'
    dv.add(rng1)
    ws.conditional_formatting.add(rng1, CellIsRule(operator='equal', formula=['"CV"'], fill=fill_cv))
    ws.conditional_formatting.add(rng1, CellIsRule(operator='equal', formula=['"AG"'], fill=fill_ag))

    if nb_jours > 15:
        rng2 = f'B12:{get_column_letter(nb_jours-14+1)}12'
        dv.add(rng2)
        ws.conditional_formatting.add(rng2, CellIsRule(operator='equal', formula=['"CV"'], fill=fill_cv))
        ws.conditional_formatting.add(rng2, CellIsRule(operator='equal', formula=['"AG"'], fill=fill_ag))

    # Nettoyage Stats Strict
    clean_zone_restricted(ws, 16, 100, 1, 10)

    curr = 16

    # TABLEAU 1 : PRESTATIONS
    headers_tab = ['Prestations', 'Jours', 'H. Totales', 'H. Nuit', 'Couleur']
    for i, h in enumerate(headers_tab):
        c = ws.cell(row=curr, column=i+1)
        c.value = h
        c.fill = fill_header
        c.font = font_header
        c.alignment = centrage
        c.border = border_thin
    curr += 1

    start_prest = curr
    for code in LISTE_PRESTATIONS:
        ws[f'A{curr}'].value = code
        ws[f'B{curr}'].value = f'=COUNTIF(B6:P6,"{code}")+COUNTIF(B12:R12,"{code}")'
        ws[f'C{curr}'].value = f'=B{curr}*8'
        ws[f'D{curr}'].value = get_night_hours_formula(code, f'B{curr}')

        for col in ['A','B','C','D','E']:
            c = ws[f'{col}{curr}']
            c.border = border_thin
            c.alignment = centrage
            c.fill = fill_none
            c.font = font_data
            if col in ['C', 'D']: c.number_format = 'General'
            if col == 'E' and code in COLORS:
                c.fill = PatternFill(start_color=COLORS[code], end_color=COLORS[code], fill_type="solid")
        curr += 1
    end_prest = curr - 1

    row_total_prest = curr
    ws[f'A{curr}'].value = "TOTAL"
    ws[f'B{curr}'].value = f'=SUM(B{start_prest}:B{end_prest})'
    ws[f'C{curr}'].value = f'=SUM(C{start_prest}:C{end_prest})'
    ws[f'D{curr}'].value = f'=SUM(D{start_prest}:D{end_prest})'
    for col in ['A','B','C','D','E']:
        c = ws[f'{col}{curr}']
        c.border = border_thin
        c.alignment = centrage
        c.fill = fill_total
        c.font = font_total
        if col in ['C', 'D']: c.number_format = 'General'
    curr += 2

    # TABLEAU 2 : REPOS
    headers_repos = ['REPOS', 'Jours', 'H. Totales', 'H. Nuit', 'Couleur']
    for i, h in enumerate(headers_repos):
        c = ws.cell(row=curr, column=i+1)
        c.value = h
        c.fill = fill_header
        c.font = font_header
        c.alignment = centrage
        c.border = border_thin
    curr += 1

    start_repos = curr
    for code in LISTE_REPOS:
        ws[f'A{curr}'].value = code
        ws[f'B{curr}'].value = f'=COUNTIF(B6:P6,"{code}")+COUNTIF(B12:R12,"{code}")'
        if code in CODES_8H: ws[f'C{curr}'].value = f'=B{curr}*8'
        else: ws[f'C{curr}'].value = 0
        ws[f'D{curr}'].value = 0
        for col in ['A','B','C','D','E']:
            c = ws[f'{col}{curr}']
            c.border = border_thin
            c.alignment = centrage
            c.fill = fill_none
            c.font = font_data
            if col in ['C', 'D']: c.number_format = 'General'
            if col == 'E' and code in COLORS:
                c.fill = PatternFill(start_color=COLORS[code], end_color=COLORS[code], fill_type="solid")
        curr += 1
    end_repos = curr - 1

    row_total_repos = curr
    ws[f'A{curr}'].value = "TOTAL"
    ws[f'B{curr}'].value = f'=SUM(B{start_repos}:B{end_repos})'
    ws[f'C{curr}'].value = f'=SUM(C{start_repos}:C{end_repos})'
    ws[f'D{curr}'].value = f'=SUM(D{start_repos}:D{end_repos})'
    for col in ['A','B','C','D','E']:
        c = ws[f'{col}{curr}']
        c.border = border_thin
        c.alignment = centrage
        c.fill = fill_total
        c.font = font_total
        if col in ['C', 'D']: c.number_format = 'General'
    curr += 2

    # TOTAL GÉNÉRAL
    ws[f'A{curr}'].value = "Total Heures"
    ws[f'C{curr}'].value = f'=C{row_total_prest}+C{row_total_repos}'
    for col in ['A', 'C']:
        c = ws[f'{col}{curr}']
        c.font = font_total
        c.border = border_thin
        c.alignment = centrage
        if col == 'C': c.number_format = 'General'

    row_total_mois = curr

    # 3. Contrôles Période
    row_ctrl = row_total_mois + 4
    ws[f'A{row_ctrl}'].value = f'CONTRÔLES PAR PÉRIODE - {mois_nom.upper()}'
    ws[f'A{row_ctrl}'].font = Font(bold=True, size=12)

    headers = ['Période', 'Heures', 'CH', 'RH', 'CV', 'RR', 'Férié Trav.', 'Statut']
    for idx, h in enumerate(headers):
        c = ws.cell(row=row_ctrl+2, column=idx+1)
        c.value = h
        c.fill = fill_header
        c.font = font_header
        c.alignment = centrage
        c.border = border_thin

    periodes_du_mois = []
    seen = set()
    for j in range(1, nb_jours+1):
        dt = datetime(ANNEE, mois_num, j)
        pid, pdeb, pfin = get_periode_pour_date(dt, periodes)
        if pid and pid not in seen:
            periodes_du_mois.append((pid, pdeb, pfin))
            seen.add(pid)

    curr_ctrl = row_ctrl + 3
    CODES_FERIE = [7101, 7102, 7111, 7112, 7121]

    for pid, deb, fin in periodes_du_mois:
        ws[f'A{curr_ctrl}'].value = f'P{pid}'
        ranges = get_excel_ranges_for_period(deb, fin, ANNEE)

        parts_h = []
        for c in CODES_8H:
            c_str = f'"{c}"' if isinstance(c, str) else str(c)
            for rng in ranges:
                parts_h.append(f'COUNTIF({rng},{c_str})')
        ws[f'B{curr_ctrl}'].value = f"=({' + '.join(parts_h)})*8" if parts_h else 0

        def make_cnt(tgt):
            p = [f'COUNTIF({r},"{tgt}")' for r in ranges]
            return "=" + "+".join(p) if p else 0

        ws[f'C{curr_ctrl}'].value = make_cnt("CH")
        ws[f'D{curr_ctrl}'].value = make_cnt("RH")
        ws[f'E{curr_ctrl}'].value = make_cnt("CV")
        ws[f'F{curr_ctrl}'].value = make_cnt("RR")

        cells_ferie = get_excel_ranges_for_ferie(deb, fin, ANNEE, jours_feries)
        if cells_ferie:
            cstr = ",".join(map(str, CODES_FERIE))
            parts = [f'SUM(COUNTIF({ref},{{{cstr}}}))' for ref in cells_ferie]
            ws[f'G{curr_ctrl}'].value = "=" + "+".join(parts)
        else:
            ws[f'G{curr_ctrl}'].value = 0

        ws[f'H{curr_ctrl}'].value = f'=IF(B{curr_ctrl}>160,"⚠ >160h",IF(C{curr_ctrl}<>4,"⚠ CH≠4",IF(D{curr_ctrl}<>4,"⚠ RH≠4",IF(E{curr_ctrl}<>1,"⚠ CV≠1",IF(AND(G{curr_ctrl}>0,F{curr_ctrl}=0),"⚠ RR manquant","✓ OK")))))'

        for col in range(1, 9):
            c = ws.cell(row=curr_ctrl, column=col)
            c.border = border_thin
            c.font = font_data
            c.alignment = centrage
            if col in range(2, 8): c.number_format = 'General'

            l = get_column_letter(col)
            ref = f'{l}{curr_ctrl}'
            if col == 2: ws.conditional_formatting.add(ref, CellIsRule(operator='greaterThan', formula=['160'], fill=alert_fill, font=alert_font))
            elif col in [3,4]: ws.conditional_formatting.add(ref, CellIsRule(operator='notEqual', formula=['4'], fill=warning_fill))
            elif col == 5: ws.conditional_formatting.add(ref, CellIsRule(operator='notEqual', formula=['1'], fill=warning_fill))
            elif col == 8:
                ws.conditional_formatting.add(ref, CellIsRule(operator='equal', formula=['"✓ OK"'], fill=ok_fill))
                ws.conditional_formatting.add(ref, CellIsRule(operator='notEqual', formula=['"✓ OK"'], fill=alert_fill, font=alert_font))
        curr_ctrl += 1

    print(f"    • {mois_nom}")

nom_final = f"planning_{ANNEE}.xlsx"
wb.save(nom_final)

print("\n" + "="*70)
print(f"✅ TERMINÉ : {nom_final}")
print("="*70 + "\n")
